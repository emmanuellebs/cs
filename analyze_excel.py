import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\emman\Downloads\Workhub_funiL CS 2.xlsx')

print("=== ABAS DISPONÍVEIS ===")
for name in wb.sheetnames:
    print(f"  • {name}")

print("\n" + "="*80)
print("ANÁLISE DETALHADA DE CADA ABA")
print("="*80 + "\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n### ABA: {sheet_name}")
    print(f"Dimensões: {ws.dimensions}\n")
    
    # Pega todas as linhas com dados
    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append(row)
    
    # Mostra tudo (ou primeiras 20 linhas se muitas)
    for i, row in enumerate(data[:20]):
        print(f"  {i}: {row}")
    
    if len(data) > 20:
        print(f"  ... ({len(data) - 20} linhas adicionais)")
